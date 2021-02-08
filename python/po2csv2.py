import pathlib
import sys
import openpyxl
import csv
import re

p = pathlib.Path('\\\\MO-QNAP01\Management\Purchasing\PurchaseOrders\\2020')
lwb = openpyxl.Workbook()
lsh = lwb.active
list_row = 1

# special process for an irregular format
oft_conf = {\
        'Default':['B21-41'],\
        '20-008' :['B21-41','B74-90','B127-144'],\
        '20-013' :['B21-39','B72-81','B127-143','B180-17'],\
        '20-011' :['B21-38','B74-89','B127-142'],\
        '20-052A':['B21-22'],\
        '20-016' :['B19-19','B28-28','B33-33'],\
        }
#ire_list = ['20-013', '20-008', '20-011']
ire_list = [k for k in oft_conf.keys()]

def irregular_format_process(x, key, list_row):
    print(x)
    wb = openpyxl.load_workbook(x, data_only=True)
    for sh in wb:
        if sh.title == 'PO':
            for ofts in oft_conf[key]:
                print('----', ofts)
#                o = re.compile(r'(\w??)(\d{1,5})-(\d{1,5})').search(ofts)
                o = re.compile(r'(\w??)(?P<START>\d{1,5})-(?P<END>\d{1,5})').search(ofts)
                for row in range(int(o.group('START')), int(o.group('END'))+1):
                    if sh['A'+str(row)].value != None:
                        values = [ \
                                   sh['A13'].value,\
                                   sh['D13'].value,\
                                   sh['M1'].value,\
                                   sh['B'+str(row)].value,\
                                   sh['F'+str(row)].value,\
                                   sh['G'+str(row)].value,\
                                   sh['I'+str(row)].value,\
                                   # Adding a summation function into a CELL in Excel file.
                                   #'=' + 'E' + str(list_row+1) + '*G' + str(list_row+1), x.name\
                                   '=E%s*G%s%s' % (str(list_row+1), str(list_row+1), x.name)\
                                 ]
                        for n, value in zip(range(1,len(values)+1), values):
                            lsh.cell(list_row, n).value = value
                    list_row += 1

    return list_row

#print(p)
xls = list(p.glob('**/**/*.xlsm'))
#print(xls)
print('file counts:', len(xls))
for x in xls:
#    print(x)
    f = re.search(r'(?P<PO#>\d\d-\d{0,3}[A-Z]?).*.xlsm', x.name)
#    print(f.group(1),'-----------------')
    if f.group('PO#') in ire_list:
        list_row = irregular_format_process(x, f.group('PO#'), list_row)
    elif x.match(r'*R211*.xlsm'):
        list_row = irregular_format_process(x, 'Default', list_row)
    else:
        pass

columns_titile = ['PO#', 'date', 'vender', 'desc', 'qty', 'unit', 'ucst', 'amt', 'file']
with open('po.csv', 'wt', encoding='utf-8_sig') as fp:
    writer = csv.writer(fp, lineterminator='\n')
    # writting colum titles
    writer.writerow(columns_titile)
    for row in lsh.rows:
        writer.writerow([col.value for col in row])

