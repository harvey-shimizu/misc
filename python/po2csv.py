import pathlib
import sys
import openpyxl
import csv
import re


p = pathlib.Path('\\\\MO-QNAP01\Management\Purchasing\PurchaseOrders\\2020')
lwb = openpyxl.Workbook()
lsh = lwb.active
list_row = 1

# special process for an irregular format
oft_conf = {\
        '20-008' :['B21-41','B74-90','B127-144'],\
        '20-013' :['B21-39','B72-81','B127-143','B180-17'],\
        '20-011' :['B21-38','B74-89','B127-142'],\
        '20-052A':['B21-22'],\
        '20-016' :['B19-19','B28-28', 'B33-33'],\
        }
#ire_list = ['20-013', '20-008', '20-011']
ire_list = [k for k in oft_conf.keys()]

def irregular_format_process(x, key, list_row):
    print(x)
    wb = openpyxl.load_workbook(x, data_only=True)
    value = oft_conf[key]
    for sh in wb:
        if sh.title == 'PO':
            for ofts in value:
                print('----', ofts)
                o = re.search(r'(\w??)(\d{0,5})-(\d{0,5})', ofts)
                for row in range(int(o.group(2)), int(o.group(3))+1):
                    colm = 1
                    lsh.cell(list_row, colm+0).value = sh['A13'].value
                    lsh.cell(list_row, colm+1).value = sh['D13'].value
                    lsh.cell(list_row, colm+2).value = sh['M1'].value
                    lsh.cell(list_row, colm+3).value = sh['B'+str(row)].value
                    lsh.cell(list_row, colm+4).value = sh['F'+str(row)].value
                    lsh.cell(list_row, colm+5).value = sh['G'+str(row)].value
                    lsh.cell(list_row, colm+6).value = sh['I'+str(row)].value
                    lsh.cell(list_row, colm+7).value = '=' + 'E' + str(list_row+1) + '*G' + str(list_row+1)
                    lsh.cell(list_row, colm+8).value = x.name
                    list_row += 1
    return list_row

#print(p)
xls = list(p.glob('**/**/*.xlsm'))
#print(xls)
print('file counts:', len(xls))
for x in xls:
#    print(x)
    f = re.search(r'(\d\d-\d{0,3}[A-Z]?).*.xlsm', x.name)
#    print(f.group(1),'-----------------')
    if f.group(1) in ire_list:
        list_row = irregular_format_process(x, f.group(1), list_row)
    elif x.match(r'*R211*.xlsm'):
        print(x)
        wb = openpyxl.load_workbook(x, data_only=True)
        for sh in wb:
            if sh.title == 'PO':
                colm = 1
                lsh.cell(list_row, colm+0).value = sh['A13'].value
                lsh.cell(list_row, colm+1).value = sh['D13'].value
                lsh.cell(list_row, colm+2).value = sh['M1'].value
                lsh.cell(list_row, colm+3).value = sh['B19'].value
                lsh.cell(list_row, colm+4).value = sh['F19'].value
                lsh.cell(list_row, colm+5).value = sh['G19'].value
                lsh.cell(list_row, colm+6).value = sh['I19'].value
                lsh.cell(list_row, colm+7).value = '=' + 'E' + str(list_row+1) + '*G' + str(list_row+1)
                lsh.cell(list_row, colm+8).value = x.name
                list_row += 1

columns_titile = ['PO#', 'date', 'vender', 'desc', 'qty', 'unit', 'ucst', 'amt', 'file']
with open('po.csv', 'wt', encoding='utf-8_sig') as fp:
    writer = csv.writer(fp, lineterminator='\n')
    # writting colum titles
    writer.writerow(columns_titile)
    for row in lsh.rows:
        writer.writerow([col.value for col in row])

